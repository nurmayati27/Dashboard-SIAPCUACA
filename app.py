# app.py
from datetime import datetime, date, time, timedelta
from urllib.parse import quote
from functools import lru_cache
import pandas as pd
import streamlit as st
import re, unicodedata
import numpy as np
import os
from difflib import SequenceMatcher
import streamlit.components.v1 as components
from supabase import create_client, Client

# ✅ letakkan helper ini di sini (setelah import, sebelum fungsi lain)
def _secret(name: str, default: str = "") -> str:
    try:
        # format nested: [SUPABASE] URL=..., ANON_KEY=...
        if "SUPABASE" in st.secrets:
            m = st.secrets["SUPABASE"]
            if name in m: 
                return m[name]
    except Exception:
        pass
    # format flat: SUPABASE_URL=..., SUPABASE_ANON_KEY=...
    return st.secrets.get(f"SUPABASE_{name}", os.environ.get(f"SUPABASE_{name}", default))

# ✅ ambil rahasia Supabase (URL & ANON_KEY)
SUPABASE_URL = _secret("URL")
SUPABASE_ANON_KEY = _secret("ANON_KEY")

# ✅ buat koneksi Supabase
sb = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

# ====== UI ======
st.set_page_config(page_title="SIAPCUACA", layout="wide")
st.title("📊 Sistem Informasi Arsip Cuaca (SIAPCUACA)")

DEBUG_SIM = True  
_sem_model = None

# ==== BOOT PROFILER: pantau waktu load di terminal ====
import time as _t
__boot_t0 = _t.perf_counter()

def BOOT(msg: str):
    # tulis progress startup ke terminal streamlit
    print(f"[BOOT +{_t.perf_counter()-__boot_t0:0.3f}s] {msg}")

# === JS: auto-resize semua <textarea> Streamlit ===
AUTOSIZE_JS = """
<script>
(function () {
  // cari pembungkus terdekat yang ngunci tinggi dan buka kuncinya
  function unlockWrappers(t) {
    // Base Web wrapper di Streamlit biasanya 2-3 tingkat di atas <textarea>
    let p = t.parentElement;
    for (let i = 0; i < 4 && p; i++) {
      // lepas height / maxHeight / overflow kalau ada
      p.style.height = 'auto';
      p.style.maxHeight = 'none';
      if (getComputedStyle(p).overflowY === 'auto' || getComputedStyle(p).overflowY === 'scroll') {
        p.style.overflowY = 'visible';
      }
      p = p.parentElement;
    }
  }

  function fit(t) {
    if (!t) return;
    unlockWrappers(t);
    t.style.overflow = 'hidden';
    t.style.height = 'auto';
    // tambah sedikit padding supaya tidak kepotong baris terakhir
    t.style.height = (t.scrollHeight + 2) + 'px';
  }

  function hookAll() {
    document.querySelectorAll('textarea').forEach(t => {
      if (t.dataset._auto) return;
      t.dataset._auto = '1';
      // pertama kali
      fit(t);
      // saat user mengetik / paste
      t.addEventListener('input', () => fit(t));
      // setelah paste besar, tunggu layout settle lalu fit lagi
      t.addEventListener('paste', () => setTimeout(() => fit(t), 0));
    });
  }

  // jalankan sekarang dan setiap ada node baru (rerender Streamlit)
  hookAll();
  new MutationObserver(hookAll).observe(document.body, { subtree: true, childList: true });
})();
</script>
"""

# Opsional indikator di sidebar TANPA memaksa load:
if DEBUG_SIM:
    st.sidebar.write(
        "🧠 Model semantik:",
        ("OK (sudah dimuat)" if _sem_model is not None else "Belum dimuat")
    )

# injeksi JS auto-resize SEKALI untuk semua <textarea>
BOOT("before inject CSS/JS")
st.markdown(AUTOSIZE_JS, unsafe_allow_html=True)
BOOT("after inject CSS/JS")

BOOT("start app.py (imports selesai)")

# === DEBUG VIEW SWITCH (sembunyikan tampilan, logika tetap jalan) ===
def dbg(msg: str):
    if DEBUG_SIM:
        st.caption(msg)

import time

class Tmr:
    def __init__(self, label): self.t0=time.perf_counter(); self.label=label
    def tick(self, msg): 
        dt = time.perf_counter()-self.t0
        if DEBUG_SIM:                           # ⟵ hanya render saat debug ON
            st.caption(f"⏱️ {self.label}: {msg} — {dt:0.3f}s")
        self._t0 = time.perf_counter()


# --- Pola bullet/penomoran: "-", "•", "1. ", "2) ", dsb.
ITEM_HDR = re.compile(r'^\s*(?:[-–•]|\d+\s*[.)])\s+')

def _split_items(text: str) -> list[str]:
    """
    Untuk PERTANYAAN: 1 baris/1 bullet = 1 item.
    Baris kosong memutus item; kalau ada bullet/penomoran di awal baris,
    label bulletnya dibuang tapi isinya tetap jadi satu item.
    """
    if not text:
        return []
    # Normalisasi newline
    text = re.sub(r'\r\n?', '\n', text)
    out, cur = [], []

    def push():
        s = '\n'.join(cur).strip()
        if s:
            out.append(s)

    for ln in text.split('\n'):
        if not ln.strip():           # baris kosong → tutup item
            if cur:
                push(); cur = []
            continue
        if ITEM_HDR.match(ln):       # bullet/nomor di awal baris → mulai item baru
            if cur:
                push(); cur = []
            ln = ITEM_HDR.sub('', ln, count=1)  # hilangkan label bullet
        cur.append(ln)

    if cur:
        push()
    return out


def _split_answers_by_paragraph(text: str) -> list[str]:
    """
    Untuk JAWABAN: pecah HANYA kalau ada paragraf kosong (double enter).
    Single enter/ bullet/ numbering tetap menyatu di satu jawaban.
    """
    if not text:
        return []
    text = re.sub(r'\r\n?', '\n', text)
    parts = re.split(r'\n\s*\n', text)   # pisah oleh 1+ baris kosong
    return [p.strip() for p in parts if p.strip()]


# ====== Ag-Grid ======
from st_aggrid import (
    AgGrid, 
    GridOptionsBuilder, 
    ColumnsAutoSizeMode, 
    GridUpdateMode,
)
try:
    from st_aggrid import JsCode
except ImportError:
    from st_aggrid.shared import JsCode  # fallback versi lama

# ====== Supabase ======
from supabase import create_client

@st.cache_data(ttl=15)
def load_data(limit: int | None = None) -> pd.DataFrame:
    # Ambil hanya kolom yang digunakan di dashboard
    q = sb.table("questions").select("id,question,answer,session_time,media,journalist,timestamp").order("timestamp", desc=True)
    if limit:
        q = q.limit(limit)
    
    data = q.execute().data or []
    df = pd.DataFrame(data)

    # Pastikan kolom wajib tetap ada (kalau DB kosong)
    for col in ["id","timestamp","journalist","media","session_time","question","answer"]:
        if col not in df.columns:
            df[col] = None

    # ✅ Konversi waktu & buat kolom tanggal turunan
    df["session_time"] = pd.to_datetime(df["session_time"], errors="coerce", utc=True)
    df["__date_tmp"]   = df["session_time"].dt.date

    return df

def _chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i+n]

@st.cache_resource
def get_sb():
    return create_client(st.secrets["SUPABASE"]["URL"], st.secrets["SUPABASE"]["ANON_KEY"])

def save_rows(rows):
    if not rows:
        return
    sb = get_sb()
    for chunk in _chunks(rows, 500):
        sb.table("questions").insert(chunk).execute()
    BOOT("supabase client siap") 

def _reset_all_inputs(max_pairs: int = 50):
    """
    Paksa semua input jadi kosong (""), lalu set tgl/jam ke default baru.
    max_pairs = jumlah maksimum q{i}/a{i} yang mau dibersihkan.
    """
    # Kosongkan Q/A mode Form biasa
    for i in range(max_pairs):
        st.session_state[f"q{i}"] = ""
        st.session_state[f"a{i}"] = ""
        st.session_state.pop(f"dbg_sem_{i}", None)  # kalau ada checkbox debug

    # Kosongkan textarea mode bulk
    st.session_state["qbulk"] = ""
    st.session_state["abulk"] = ""

    # Kosongkan identitas sesi (TERMASUK “N/A” yang kamu tulis)
    st.session_state["journalist"] = ""
    st.session_state["media"] = ""

    # Reset tanggal & jam (date_input/time_input tidak bisa benar-benar kosong)
    st.session_state["tgl_wawancara"] = date.today()
    st.session_state["jam_wawancara"] = datetime.now().replace(second=0, microsecond=0).time()

    # (opsional) balik ke Form biasa kalau radionya punya key "mode"
    # st.session_state["mode"] = "Form biasa"

    
def delete_rows_by_ids(ids: list[int]) -> bool:
    if not ids:
        return False
    try:
        res = sb.table("questions").delete().in_("id", ids).execute()
        return bool(res.data)
    except Exception as e:
        st.error(f"Gagal menghapus: {e}")
        return False

def query_questions(qkey: str = "", media: str = "", journalist: str = "",
                    limit: int = 200, offset: int = 0) -> pd.DataFrame:
    q = (
        sb.table("questions")
          .select("id,question,answer,session_time,media,journalist,timestamp")
          .order("session_time", desc=True)   # ← tanpa nulls=
          .order("timestamp",   desc=True)    # ← fallback
    )

    if media:
        q = q.ilike("media", f"%{media}%")
    if journalist:
        q = q.ilike("journalist", f"%{journalist}%")
    if qkey:
        q = q.or_("question.ilike.%{0}%,answer.ilike.%{0}%,journalist.ilike.%{0}%,media.ilike.%{0}%".format(qkey))

    q = q.range(offset, offset + limit - 1)
    data = q.execute().data or []
    return pd.DataFrame(data)

def _pick_days_window(q: str) -> int:
    ql = (q or "").lower()
    # istilah yang cenderung butuh data terbaru
    recent_terms = (
        "prakiraan", "prediksi", "besok", "lusa",
        "pekan depan", "minggu depan", "prospek", "3 hari", "7 hari"
    )
    return 90 if any(t in ql for t in recent_terms) else 365*5

@st.cache_data(ttl=60)  # cache 60 detik per (q, topk, days)
def find_similar_rpc(q: str, topk: int = 20, days_window: int = 365) -> pd.DataFrame:
    if not q or not q.strip():
        return pd.DataFrame()
    payload = {"q": q, "n": topk, "days_window": days_window}
    res = sb.rpc("find_similar_questions", payload).execute()
    df = pd.DataFrame(res.data or [])
    if not df.empty:
        df["session_time"] = pd.to_datetime(df["session_time"], errors="coerce", utc=True)
        df["__date_tmp"] = df["session_time"].dt.date
        # normalisasi skor 0..1 (opsional, biar enak dipakai UI)
        df["sim"] = df["sim"].fillna(0).clip(0, 1)
    return df

def _rpc_candidates(q: str, topn: int = 300, days_window: int = 3650) -> pd.DataFrame:
    """
    Ambil kandidat lexical dari fungsi Postgres (GIN + trigram) di SELURUH arsip.
    BUKAN pembatasan tanggal; days_window besar biar longgar.
    """
    try:
        df = find_similar_rpc(q, topk=topn, days_window=days_window)
        # df kolom: id, question, answer, session_time, __date_tmp, sim (versi DB)
        if df.empty: 
            return df
        df = df.rename(columns={"__date_tmp":"__date", "sim":"sim_db"})
        return df[["id","question","answer","__date","sim_db"]].copy()
    except Exception:
        return pd.DataFrame()


# --- Ambang skor final (0..1). 0.76 pas untuk parafrase ringan; atur 0.72–0.80 sesuai selera.
SIM_THRESHOLD = 0.52    # final hybrid threshold (coba 0.78–0.82)
SEM_PREFILTER = 0.35     # minimal skor SEMANTIK agar lolos ke tahap akhir (0.55–0.62 oke)
TOPK_SEM = 40            # ambil top-K kandidat paling semantik untuk dihitung hybrid
# --- (A) Komponen lexical ---
try:
    from rapidfuzz import fuzz
    _use_rapidfuzz = True
except Exception:
    _use_rapidfuzz = False

_STOPWORDS = {
    # tanya/preposisi umum
    "apa","apakah","bagaimana","mengapa","kenapa","kepada","untuk","pada","di","ke","dari",
    "dengan","dan","atau","yang","ini","itu",
    # sangat umum (opsional – boleh hapus kalau terasa terlalu ketat)
    "indonesia","wilayah","daerah"
}

def _basic_clean(s: str) -> str:
    s = unicodedata.normalize("NFKC", s or "").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # normalisasi ejaan ringan
    s = s.replace("himbau", "imbau").replace("prakirakan","prakiraan")
    return s

def _content_tokens(s: str) -> set[str]:
    toks = [t for t in _basic_clean(s).split() if len(t) > 1 and t not in _STOPWORDS]
    # kecilkan variasi ringan
    repl = {"imbauan":"imbau","mengimbau":"imbau","diimbau":"imbau"}
    return {repl.get(t, t) for t in toks}

def _lexical_sim(a: str, b: str) -> float:
    """Gabungan RapidFuzz token_sort (string-level) + Jaccard token konten."""
    na, nb = _basic_clean(a), _basic_clean(b)
    if _use_rapidfuzz:
        rf = fuzz.token_sort_ratio(na, nb) / 100.0
    else:
        rf = SequenceMatcher(None, na, nb).ratio()

    ta, tb = _content_tokens(a), _content_tokens(b)
    if not ta or not tb:
        jac = 0.0
    else:
        jac = len(ta & tb) / len(ta | tb)

    # utamakan Jaccard (fokus isi), rf sebagai pelengkap
    return 0.6 * jac + 0.4 * rf

# --- (B) Komponen semantic (makna) ---
# Model multilingual yang paham bahasa Indonesia (±250 MB)
@st.cache_resource(show_spinner=False)
def _load_semantic_model():
    # (opsional) pastikan folder cache jelas
    os.environ.setdefault("SENTENCE_TRANSFORMERS_HOME", ".cache/st_models")
    from sentence_transformers import SentenceTransformer  # import BERAT di dalam fungsi cached
    return SentenceTransformer("distiluse-base-multilingual-cased-v2", device="cpu")

def _ensure_model():
    """Ambil model dari cache. Tidak pakai global supaya aman di rerun."""
    try:
        return _load_semantic_model()
    except Exception:
        return None  # fallback jika paket belum terpasang

# Jika kamu punya fungsi embed yang cache per teks, pastikan normalisasi aktif
@st.cache_data(show_spinner=False)
def _cached_embed(text: str) -> np.ndarray | None:
    m = _ensure_model()
    if m is None or not text:
        return None
    # penting: normalisasi -> supaya cosine == dot product
    vec = m.encode([text], convert_to_numpy=True, normalize_embeddings=True, show_progress_bar=False)[0]
    return vec.astype("float32")

def _semantic_sim(a: str, b: str) -> float:
    """Kemiripan cosine tanpa torch: cukup NumPy (karena vektor sudah dinormalisasi)."""
    if not a or not b:
        return 0.0
    ea = _cached_embed(a)
    eb = _cached_embed(b)
    if ea is None or eb is None:
        return 0.0
    # karena sudah dinormalisasi, cosine == dot product
    return float(np.dot(ea, eb))

# --- (C) Skor HYBRID akhir ---
def _hybrid_sim(a: str, b: str) -> float:
    sem = _semantic_sim(a, b)      # makna
    lex = _lexical_sim(a, b)       # kata
    return 0.80 * sem + 0.20 * lex # PORSI SEMANTIK DIBESARKAN

@st.cache_data(ttl=60, show_spinner=False)
def _get_all_questions_for_scope(limit: int = 5000) -> pd.DataFrame:
    df = query_questions(qkey="", media="", journalist="", limit=limit, offset=0).copy()
    # tentukan kolom waktu
    tcol = None
    for c in ("session_time", "timestamp_tz", "timestamp"):
        if c in df.columns:
            tcol = c; break
    if tcol and not df.empty:
        df["__ts"] = pd.to_datetime(df[tcol], errors="coerce")
        df["__date"] = df["__ts"].dt.date
        iso = df["__ts"].dt.isocalendar()
        df["__iso_year"] = iso.year.astype("Int64")
        df["__iso_week"] = iso.week.astype("Int64")
    else:
        df["__date"] = pd.NaT
        df["__iso_year"] = pd.NA
        df["__iso_week"] = pd.NA
    return df

# ==== Tambahan aturan untuk menekan false positive ====
STRICT_SEM = 0.58     # kalau semantik setinggi ini, lolos walau Jaccard kecil
MIN_JACCARD = 0.12    # kalau semantik biasa-biasa, wajib Jaccard ≥ ini
FALLBACK_MIN_SEM = 0.50
FALLBACK_MIN_JAC = 0.08

def _jaccard_content_only(a: str, b: str) -> float:
    ta, tb = _content_tokens(a), _content_tokens(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)

@lru_cache(maxsize=10000)
def _cached_embed(text: str):
    if not text:
        return None
    model = _ensure_model()
    if model is None:
        return None
    return model.encode(text, normalize_embeddings=True)


def _corpus_signature(df: pd.DataFrame) -> str:
    """Penanda perubahan korpus sederhana agar cache invalid saat data baru masuk."""
    if df.empty:
        return "0-0"
    n = len(df)
    mx = int(pd.to_numeric(df["id"], errors="coerce").fillna(0).max())
    return f"{n}-{mx}"

@st.cache_resource(show_spinner=False)
def _build_semantic_matrix(sig: str):
    """
    Bangun matriks embedding untuk SELURUH korpus (tanpa batas baris).
    Dipanggil ulang otomatis jika signature berubah.
    """
    base = _get_all_questions_for_scope(limit=50000)  # full pull
    base = base.dropna(subset=["id","question"]).copy()
    base["id"] = pd.to_numeric(base["id"], errors="coerce").astype("Int64")
    # tanggal untuk recency
    if "__date" not in base.columns:
        ts = pd.to_datetime(base.get("session_time"), errors="coerce")
        base["__date"] = ts.dt.date

    # batch encode → jauh lebih cepat daripada encode per‐baris
    model = _ensure_model()
    qs = base["question"].astype(str).tolist()
    if model is None or not qs:
        emb = np.zeros((0, 1), dtype="float32")
    else:
        emb = model.encode(
            qs, convert_to_numpy=True, normalize_embeddings=True,
            batch_size=128, show_progress_bar=False
        ).astype("float32")
    # hanya kolom yang diperlukan lanjut
    view = base[["id","question","answer","__date"]].reset_index(drop=True)
    return view, emb



# >>> Tambahkan fungsi ini di bawah query_questions() atau delete_rows_by_ids()


def update_row_by_id(row_id: int, new_q: str, new_a: str) -> bool:
    """
    Update kolom 'question' dan 'answer' di tabel 'questions' berdasarkan id.
    Kembalikan True jika berhasil.
    """
    try:
        res = (
            sb.table("questions")
              .update({"question": new_q, "answer": new_a})
              .eq("id", row_id)
              .execute()
        )
        if res.data:  # Supabase balikin list of dict
            return True
        else:
            st.warning(f"Tidak ada baris yang diperbarui untuk id={row_id}")
            return False
    except Exception as e:
        st.error(f"Gagal update id={row_id}: {e}")
        return False


def _fill_bulk_answer(idx: int, ans_text: str, total_q: int):
    """Callback untuk mengisi jawaban ke indeks 'idx' pada textarea 'abulk'."""
    cur = (st.session_state.get("abulk", "") or "")
    lst = _split_answers_by_paragraph(cur)

    # pastikan panjang list jawaban >= jumlah pertanyaan
    if len(lst) < total_q:
        lst += [""] * (total_q - len(lst))

    lst[idx] = (ans_text or "").strip()
    st.session_state["abulk"] = "\n\n".join(lst)

    # opsional: kirim pesan sukses agar bisa ditampilkan setelah rerun
    st.session_state["__last_filled_msg"] = f"Jawaban dimasukkan ke Q#{idx+1}"

# ============================================================
# GLOBAL SEMANTIC INDEX (FAISS) — cepat & tanpa batas jendela hari
# ============================================================
import os, numpy as np, pandas as pd, streamlit as st
from datetime import date
os.environ.setdefault("SENTENCE_TRANSFORMERS_HOME", ".cache/st_models")

@st.cache_resource(show_spinner=False)
def _sim_model():
    """Load model sekali (kecil & cepat, 384 dim)."""
    from sentence_transformers import SentenceTransformer
    return SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2", device="cpu")

def _encode_batch(texts: list[str]) -> np.ndarray:
    m = _sim_model()
    return m.encode(texts, convert_to_numpy=True, normalize_embeddings=True, show_progress_bar=False).astype("float32")

def _fmt_dt(v) -> str:
    try:
        return pd.to_datetime(v, errors="coerce").strftime("%Y-%m-%d %H:%M")
    except Exception:
        return "" if (pd.isna(v) or v is None) else str(v)

# Ambil SELURUH korpus (pakai fungsi kamu)
@st.cache_data(ttl=300, show_spinner=False)
def _load_corpus_df() -> pd.DataFrame:
    """Ambil seluruh data pertanyaan secara bertahap (chunked) agar tidak timeout."""
    all_chunks = []
    offset = 0
    chunk = 5000
    max_rows = 200000  # batas aman biar gak overload server

    while True:
        try:
            df = query_questions(
                qkey="", media="", journalist="",
                limit=chunk, offset=offset
            )
        except Exception as e:
            st.warning(f"Gagal ambil batch offset={offset}: {e}")
            break

        if df is None or df.empty:
            break

        all_chunks.append(df)
        n = len(df)
        offset += n
        if n < chunk or offset >= max_rows:
            break
        time.sleep(0.3)  # jeda kecil antar batch

    if not all_chunks:
        return pd.DataFrame(columns=["id", "question", "answer", "session_time"])

    df = pd.concat(all_chunks, ignore_index=True)
    for c in ("id", "question", "answer", "session_time", "timestamp_tz", "timestamp"):
        if c not in df.columns:
            df[c] = pd.NA

    df["question"] = df["question"].fillna("").astype(str)
    df = df[df["question"].str.strip() != ""].copy().reset_index(drop=True)
    return df

# signature untuk invalidasi cache saat data berubah (murah)
def _corpus_sig(df: pd.DataFrame) -> str:
    if df.empty: return "0:0"
    n = len(df); mx = int(pd.to_numeric(df["id"], errors="coerce").fillna(0).max())
    return f"{n}:{mx}"

@st.cache_data(show_spinner=False)
def _corpus_embeddings(sig: str, df: pd.DataFrame) -> dict:
    ids   = df["id"].to_numpy()
    texts = df["question"].tolist()
    vecs  = _encode_batch(texts) if texts else np.zeros((0,384), dtype="float32")
    return {"ids": ids, "vecs": vecs, "df": df}

@st.cache_resource(show_spinner=True)
def _faiss_index(sig: str, vecs: np.ndarray):
    import faiss
    index = faiss.IndexFlatIP(vecs.shape[1])  # inner product = cosine (karena normalized)
    if len(vecs): index.add(vecs)
    return index

def find_similar_global(q_text: str, top_k: int = 3) -> list[dict]:
    """Cari mirip untuk 1 teks terhadap seluruh arsip (tanpa batas hari)."""
    q_text = (q_text or "").strip()
    if not q_text: return []
    df = _load_corpus_df()
    if df.empty: return []
    sig   = _corpus_sig(df)
    corp  = _corpus_embeddings(sig, df)
    index = _faiss_index(sig, corp["vecs"])

    qv = _encode_batch([q_text])
    scores, idxs = index.search(qv, top_k)
    out = []
    for j, i in enumerate(idxs[0]):
        if i < 0: continue
        row = corp["df"].iloc[int(i)]
        out.append({
            "id": int(row["id"]) if pd.notna(row["id"]) else 0,
            "score": float(scores[0][j]),
            "question": (row.get("question") or "").strip(),
            "answer":   (row.get("answer") or "").strip(),
            "sesi_dt":  _fmt_dt(row.get("session_time") or row.get("timestamp_tz") or row.get("timestamp")),
        })
    return out

def find_similar_global_bulk(texts: list[str], top_k: int = 1) -> list[list[dict]]:
    """Batch: lebih cepat untuk banyak pertanyaan sekaligus."""
    texts = [(t or "").strip() for t in texts]
    if not any(texts): return [[] for _ in texts]
    df = _load_corpus_df()
    if df.empty: return [[] for _ in texts]
    sig   = _corpus_sig(df)
    corp  = _corpus_embeddings(sig, df)
    index = _faiss_index(sig, corp["vecs"])

    qs = _encode_batch(texts)
    import faiss
    scores, idxs = index.search(qs, top_k)
    results: list[list[dict]] = []
    for m in range(len(texts)):
        bucket: list[dict] = []
        for j, i in enumerate(idxs[m]):
            if i < 0: continue
            row = corp["df"].iloc[int(i)]
            bucket.append({
                "id": int(row["id"]) if pd.notna(row["id"]) else 0,
                "score": float(scores[m][j]),
                "question": (row.get("question") or "").strip(),
                "answer":   (row.get("answer") or "").strip(),
                "sesi_dt":  _fmt_dt(row.get("session_time") or row.get("timestamp_tz") or row.get("timestamp")),
            })
        results.append(bucket)
    return results


# ============================================================
# Sidebar opsional: tombol preload manual
# ============================================================
with st.sidebar.expander("⚙️ Cek Kemiripan Pertanyaan Disini", expanded=False):
    if st.button("Preload Model"):
        from time import time
        t0 = time()
        _ = _load_semantic_model()  # hanya load model
        st.success(f"✅ Model siap dalam {time()-t0:.1f} detik")

tabs = st.tabs(["📝 Input", "📋 Rekap", "📤 Export"])
BOOT("tabs/layout siap")

def _normalize_text(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r" [^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    # normalisasi ejaan umum BI
    s = s.replace("himbau", "imbau")   # himbauan → imbauan
    s = s.replace("di  ", "di ")       # spasi ganda umum
    return s

# ==== RESET HANDLER (jalan sebelum widget dibuat) ====
def _do_full_reset(max_pairs: int = 50):
    # kosongkan semua q{i}/a{i}
    for i in range(max_pairs):
        st.session_state[f"q{i}"] = ""
        st.session_state[f"a{i}"] = ""
        st.session_state.pop(f"dbg_sem_{i}", None)

    # kosongkan textarea bulk
    st.session_state["qbulk"] = ""
    st.session_state["abulk"] = ""

    # kosongkan identitas
    st.session_state["journalist"] = ""
    st.session_state["media"] = ""

    # reset tanggal & jam (widget date/time tidak bisa null)
    st.session_state["tgl_wawancara"] = date.today()
    st.session_state["jam_wawancara"] = datetime.now().replace(second=0, microsecond=0).time()

    # (opsional) kembalikan mode
    # st.session_state["mode"] = "Form biasa"

# Jalankan reset jika flag di-set pada run sebelumnya
_reset_req = st.session_state.pop("__RESET_FORM__", False)
if _reset_req:
    _do_full_reset(max_pairs=50)


# =================== TAB INPUT ===================
with tabs[0]:
    st.subheader("Input Pertanyaan Baru")

    c1, c2 = st.columns(2)
    with c1:
        journalist = st.text_input("Nama Wartawan", key="journalist")
        media      = st.text_input("Media", key="media")
    with c2:
        # default sekali saja
        if "tgl_wawancara" not in st.session_state:
            st.session_state.tgl_wawancara = date.today()
        if "jam_wawancara" not in st.session_state:
            st.session_state.jam_wawancara = datetime.now().replace(second=0, microsecond=0).time()

        d = st.date_input("Tanggal Wawancara", value=st.session_state.tgl_wawancara, key="tgl_wawancara")
        t = st.time_input("Jam Wawancara", value=st.session_state.jam_wawancara, step=60, key="jam_wawancara")
        session_time = datetime.combine(d, t)
        _get_all_questions_for_scope.clear()  # refresh kandidat saat tanggal diubah
        
    st.markdown("### Mode Input")
    mode = st.radio("Pilih mode", ["Form biasa", "Tempel banyak (multi Q&A)"], horizontal=True)

    rows_to_save = []

    # ----------------- MODE: FORM BIASA -----------------
    if mode == "Form biasa":
        n = st.number_input("Jumlah pertanyaan", min_value=1, max_value=20, value=1, step=1)
        for i in range(n):
            st.markdown(f"**Q&A #{i+1}**")
            q = st.text_area(f"Pertanyaan #{i+1}", key=f"q{i}")
            

            # === INSERT 2: Warning mirip (Form biasa) ===
            if q.strip():
                # Gunakan hasil yang sama untuk UI (tidak panggil ulang)
                ms = find_similar_global(q, top_k=2)
                st.caption("🧩 Debug skor kemiripan:")
                SHOW_DEBUG = False
                if SHOW_DEBUG:
                    if DEBUG_SIM and st.checkbox("🔍 Lihat 10 kandidat semantik teratas (debug)", value=False, key=f"dbg_sem_{i}"):
                        _df = _get_all_questions_for_scope()[["id","question"]].copy()
                        _df["sem"] = _df["question"].apply(lambda s: _semantic_sim(q, s))
                        st.dataframe(_df.nlargest(10, "sem")[["sem","id","question"]], use_container_width=True)

                if ms:
                    for m in ms:
                        with st.container(border=True):
                            st.markdown(f"⚠️ **Mirip dengan ID #{m['id']}** (skor `{m['score']}`) · {m['sesi_dt']}")
                            st.markdown(f"**Q terdahulu:** {m['question']}")

                            ans_txt = (m.get("answer") or "").strip()
                            if ans_txt:
                                # preview 200 karakter agar ringkas
                                preview = (ans_txt[:200] + "…") if len(ans_txt) > 200 else ans_txt
                                st.markdown(f"**Preview jawaban:** {preview}")

                                with st.expander(f"⬇️ Lihat jawaban lengkap ID #{m['id']}"):
                                    st.markdown(ans_txt)

                                if st.button(f"Gunakan jawaban ID #{m['id']}", key=f"use_ans_form_{i}_{m['id']}"):
                                    st.session_state[f'a{i}'] = ans_txt
                                    st.toast(f"Jawaban dari ID #{m['id']} dimasukkan ke 'Jawaban #{i+1}'.", icon="✅")
                                    st.rerun()
                            else:
                                st.caption("_Data mirip belum memiliki jawaban._")

            a = st.text_area(f"Jawaban #{i+1}", key=f"a{i}")

        # Tombol simpan KHUSUS Form biasa (selalu tampil)
        if st.button("Simpan", key="btn_save_form"):
            rows_to_save = []
            for i in range(n):
                q = (st.session_state.get(f"q{i}", "") or "").strip()
                a = (st.session_state.get(f"a{i}", "") or "").strip()
                if q and a:
                    rows_to_save.append({
                        "timestamp": datetime.now().isoformat(timespec="seconds"),
                        "journalist": journalist,
                        "media": media,
                        "session_time": session_time.isoformat(timespec="seconds"),
                        "question": q,
                        "answer": a,
                    })

            if journalist.strip() and media.strip() and rows_to_save:
                with st.spinner("Menyimpan..."):
                    save_rows(rows_to_save)  # <-- pastikan ini bulk insert

                st.success(f"✅ Berhasil simpan {len(rows_to_save)} baris untuk 1 sesi wawancara.")

                # Minta reset pada run berikutnya lalu rerun
                st.session_state["__RESET_FORM__"] = True
                st.toast("🧹 Form berhasil direset.", icon="✅")
                st.rerun()
               
            else:
                st.warning("Lengkapi identitas sesi dan isi Q&A yang berpasangan.")

    # ----------------- MODE: TEMPEL BANYAK -----------------
    elif mode == "Tempel banyak (multi Q&A)":
        cqa1, cqa2 = st.columns(2)

        with cqa1:
            q_bulk = st.text_area(
                "📥 Pertanyaan (bisa dari bullet/list)",
                key="qbulk",
                height=180,  # tinggi awal kecil; JS akan membesarkan otomatis
            )
        with cqa2:
            a_bulk = st.text_area(
                "📥 Jawaban (urutan sama)",
                key="abulk",
                height=180,
            )

        # ✅ Tampilkan notifikasi jika ada jawaban yang baru diisi dari tombol "Gunakan jawaban ..."
        msg = st.session_state.pop("__last_filled_msg", "")
        if msg:
            st.toast(msg, icon="✅")


        # --- proses parsing & preview (pakai helper di bawah; top-level def!) ---
        qs  = _split_items(st.session_state.get("qbulk", ""))
        ans = _split_answers_by_paragraph(st.session_state.get("abulk", ""))


        # === CEK KEMIRIPAN (BULK) — pakai FAISS batch ===
        bulk_matches_lists = find_similar_global_bulk(qs, top_k=1)
        bulk_matches = [lst[0] if lst else None for lst in bulk_matches_lists]
        any_warn = any(m is not None for m in bulk_matches)

        # (Opsional) Tampilkan ringkasan warning di atas preview
        if any_warn:
            st.caption("🔎 Deteksi kemiripan (semua item yang ditempel)")
            for idx, m in enumerate(bulk_matches):
                if not m:
                    continue
                with st.container(border=True):
                    score_txt = f"{m['score']:.2f}".rstrip("0").rstrip(".")
                    st.markdown(
                        f"🟡 **Q#{idx+1}** mirip dengan **ID #{m['id']}** "
                        f"(skor `{score_txt}`) · {m['sesi_dt']}"
                    )
                    st.markdown(f"**Q terdahulu:** {m['question']}")
                
                ans_txt = (m.get("answer") or "").strip()
                if ans_txt:
                    preview = (ans_txt[:200] + "…") if len(ans_txt) > 200 else ans_txt
                    st.markdown(f"**Preview jawaban:** {preview}")
                    with st.expander(f"⬇️ Lihat jawaban lengkap ID #{m['id']}"):
                        st.markdown(ans_txt)

                if m["answer"]:
                    # tombol untuk memasukkan jawaban ke posisi Q#idx+1 di textarea 'Jawaban (urutan sama)'
                    st.button(
                        f"Gunakan jawaban ID #{m['id']} untuk Q#{idx+1}",
                        key=f"use_ans_bulk_{idx}_{m['id']}",
                        on_click=_fill_bulk_answer,
                        args=(idx, m["answer"], len(qs))
                    )
                    st.caption("💬 Jawaban sudah ada – rujuk untuk konsistensi.")
                else:
                    st.caption("_Data mirip belum memiliki jawaban._")

        else:
            st.caption("Tidak ada kemiripan yang berarti pada batch ini (di atas ambang).")


        if qs or ans:
            st.markdown("**Pratinjau pasangan Q&A**")
            m = min(len(qs), len(ans))
            # Tambahkan kolom indikator kemiripan untuk pratinjau
            mirip_id = []
            mirip_score = []
            for i in range(m):
                mm = bulk_matches[i] if i < len(bulk_matches) else None
                if mm:
                    mirip_id.append(mm["id"])
                    mirip_score.append(round(float(mm["score"]), 2))
                else:
                    mirip_id.append("")
                    mirip_score.append("")

            df_preview = pd.DataFrame({
                "pertanyaan": qs[:m],
                "jawaban": ans[:m],
            })
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            if len(qs) != len(ans):
                st.warning(f"Jumlah tidak sama: pertanyaan {len(qs)} vs jawaban {len(ans)}. "
                           f"Hanya {m} pasangan pertama yang disimpan.")

            # Tombol simpan KHUSUS Tempel banyak (selalu tampil)
            if st.button("Simpan", key="btn_save_bulk"):
                rows_to_save = []
                qs = _split_items(st.session_state.get("qbulk", ""))
                ans = _split_items(st.session_state.get("abulk", ""))
                m = min(len(qs), len(ans))
                for q, a in zip(qs[:m], ans[:m]):
                    if q.strip() and a.strip():
                        rows_to_save.append({
                            "timestamp": datetime.now().isoformat(timespec="seconds"),
                            "journalist": journalist,
                            "media": media,
                            "session_time": session_time.isoformat(timespec="seconds"),
                            "question": q.strip(),
                            "answer": a.strip(),
                        })

                if journalist.strip() and media.strip() and rows_to_save:
                    with st.spinner("Menyimpan..."):
                        save_rows(rows_to_save)  # <-- pastikan bulk insert

                    st.success(f"✅ Berhasil simpan {len(rows_to_save)} baris untuk 1 sesi wawancara.")

                    st.session_state["__RESET_FORM__"] = True
                    st.toast("🧹 Form berhasil direset.", icon="✅")
                    st.rerun()

                else:
                    st.warning("Lengkapi identitas sesi dan pastikan ada pasangan Q&A yang valid.")


# ============== TAB REKAP (dikelompokkan per sesi) ==============
def render_rekap():
    st.subheader("📊 Rekap")
    st.caption("checkpoint: masuk Tab Rekap")

    # --- state awal gate ---
    st.session_state.setdefault("rekap_loaded", False)

    c1, c2 = st.columns(2)
    if c1.button("📥 Muat data rekap", key="btn_load_rekap"):
        st.session_state["rekap_loaded"] = True
    if c2.button("🔄 Reset rekap", key="btn_reset_rekap"):
        st.session_state["rekap_loaded"] = False

    # stop HANYA dari fungsi ini (bukan seluruh app)
    if not st.session_state["rekap_loaded"]:
        st.info("Klik **Muat data rekap** untuk menampilkan tabel rekap.")
        return

    # ---------- Filter & data ----------
    qkey = st.text_input("Cari (kata kunci)", key="grp_q", value="")

    df_raw = query_questions(qkey=qkey, media="", journalist="", limit=10000, offset=0)

    st.caption(f"DEBUG: jumlah data dari Supabase = {len(df_raw)} baris")

    # DEBUG aman: cari tmin/tmax dari gabungan kolom waktu yang ada
    t = pd.Series(pd.NaT, index=df_raw.index)
    for c in ["session_time", "timestamp_tz", "timestamp"]:
        if c in df_raw.columns:
            t = t.fillna(pd.to_datetime(df_raw[c], errors="coerce"))

    st.caption(f"DEBUG: df_raw tmin={t.min()} tmax={t.max()} n={len(df_raw)}")

    import datetime as dt

    # ===== Normalisasi waktu & siapkan df_f =====
    df_f = df_raw.copy()

    # kolom waktu gabungan: t = coalesce(session_time, timestamp_tz, timestamp)
    df_f["t"] = pd.NaT
    for col in ["session_time", "timestamp_tz", "timestamp"]:
        if col in df_f.columns:
            df_f[col] = pd.to_datetime(df_f[col], errors="coerce")
            df_f["t"] = df_f["t"].fillna(df_f[col])

    df_f["year"] = df_f["t"].dt.year

    with st.expander("🔎 Filter waktu", expanded=True):
        colY, colA, colB, colC = st.columns([1,1,1,1])

        # Opsi Tahun dari data
        tahun_tersedia = sorted([int(y) for y in df_f["year"].dropna().unique()], reverse=True)
        pilih_tahun = colY.multiselect("Tahun", tahun_tersedia, default=[])

        # 90 hari terakhir (OFF by default)
        use_last_90 = colA.checkbox("90 hari terakhir", value=False)

        # Rentang default = min..max dari data (fallback aman)
        tmin = df_f["t"].min()
        tmax = df_f["t"].max()
        if pd.isna(tmin): tmin = pd.Timestamp("2000-01-01")
        if pd.isna(tmax): tmax = pd.Timestamp.today()

        start_date = colB.date_input("Dari", tmin.date(), key="rekap_from")
        end_date   = colC.date_input("Sampai", tmax.date(), key="rekap_to")

        # ==== Terapkan filter ====
        filtered = df_f.dropna(subset=["t"]).copy()

        if use_last_90:
            batas = pd.Timestamp.today().normalize() - pd.Timedelta(days=90)
            filtered = filtered[filtered["t"] >= batas]
            range_info = (batas.date(), dt.date.today())

        elif pilih_tahun:
            filtered = filtered[filtered["year"].isin(pilih_tahun)]
            range_info = (
                filtered["t"].min().date() if not filtered.empty else start_date,
                filtered["t"].max().date() if not filtered.empty else end_date,
            )

        else:
            mask = (filtered["t"].dt.date >= start_date) & (filtered["t"].dt.date <= end_date)
            filtered = filtered.loc[mask].copy()
            range_info = (start_date, end_date)

        # Urutkan final: t desc, lalu timestamp desc (kalau ada)
        filtered = filtered.sort_values(
            by=["t", "timestamp"] if "timestamp" in filtered.columns else ["t"],
            ascending=[False, False] if "timestamp" in filtered.columns else [False],
            na_position="last",
            kind="mergesort",
        ).reset_index(drop=True)

        df_f = filtered

        st.caption(
            f"Filter aktif → Tahun: {', '.join(map(str, pilih_tahun)) if pilih_tahun else 'semua'}; "
            f"Tanggal: {range_info[0]} s.d. {range_info[1]}."
        )

    if df_f.empty:
        st.info("Belum ada data atau tidak ada yang cocok dengan filter.")
        return  # ← aman, karena MASIH di dalam fungsi

    # ---------- Tentukan & format waktu (pakai df_f hasil filter) ----------
    df_fmt = df_f.copy()   # <- ganti nama saja biar jelas

    # (baru) selalu gunakan kolom gabungan 't'
    df_fmt = df_f.copy()  # pakai hasil filter
    ts = pd.to_datetime(df_fmt["t"], errors="coerce")
    df_fmt["tanggal"] = ts.dt.strftime("%Y-%m-%d").fillna("")
    df_fmt["jam"]     = ts.dt.strftime("%H:%M").fillna("")


    # Label header sesi tetap seperti punyamu
    df_fmt["sesi"] = (
        df_fmt["tanggal"].astype(str).str.strip() + " | " +
        df_fmt["jam"].astype(str).str.strip()      + " | " +
        df_fmt.get("media", "").astype(str).str.strip() + " | " +
        df_fmt.get("journalist", "").astype(str).str.strip()
    )

    # Siapkan dataframe untuk grid — struktur TIDAK diubah
    df_view = (
        df_fmt[["id", "sesi", "question", "answer"]]
        .fillna("")
        .sort_values(["sesi", "id"], ascending=[False, False])
        .reset_index(drop=True)
    )
    
    # Tambahkan nomor urut per sesi (mulai dari 1 lagi di tiap grup sesi)
    df_view["No"] = df_view.groupby("sesi").cumcount() + 1

    # Susun ulang kolom agar No tampil di paling kiri
    df_view = df_view[["No", "id", "sesi", "question", "answer"]]

    # ====== BAGIAN SHARE WA (setelah df_view dibuat) ======
    from urllib.parse import quote

    st.markdown("### 📲 Share jawaban ke WhatsApp (per sesi)")

    # Pastikan kolom 'sesi' ada (fallback sederhana kalau belum ada)
    if "sesi" not in df_view.columns:
        ts = pd.to_datetime(df_view.get("timestamp", pd.NaT), errors="coerce")
        tanggal = ts.dt.strftime("%Y-%m-%d").fillna("")
        jam     = ts.dt.strftime("%H:%M").fillna("")
        sesi_fallback = (
            tanggal.astype(str).str.strip() + " | " +
            jam.astype(str).str.strip() + " | " +
            df_view.get("media", "").astype(str).str.strip() + " | " +
            df_view.get("journalist", "").astype(str).str.strip()
        )
        df_view["sesi"] = sesi_fallback

    # === Opsi penyusunan teks ===
    c_opt1, c_opt2, c_opt3 = st.columns([2, 1, 1])

    with c_opt1:
        # daftar sesi (terbaru di atas)
        sesi_list = sorted(df_view["sesi"].dropna().unique().tolist(), reverse=True)
        sesi_pick = st.selectbox("Pilih sesi wawancara", sesi_list)

    with c_opt2:
        include_header  = st.checkbox("Sertakan header sesi", value=True)
        include_number  = st.checkbox("Nomori Q/A", value=True)

    with c_opt3:
        include_question = st.checkbox("Sertakan pertanyaan", value=True)
        blank_between    = st.checkbox("Baris kosong antar Q/A", value=True)

    if sesi_pick: 
    # Ambil baris untuk sesi yang dipilih
        rows = (
            df_view.loc[df_view["sesi"] == sesi_pick, ["id", "question", "answer"]]
                  .sort_values(["id"], ascending=True)
                  .reset_index(drop=True)
        )

        # Susun item per Q&A
        items = []
        for i, r in rows.iterrows():
            num = f"{i+1}. " if include_number else ""
            q   = (r.get("question") or "").strip()
            a   = (r.get("answer") or "").strip()

            if include_question:
                # Format: nomor di depan Q, A di baris berikut
                # contoh:
                # 1. Q: ....
                #    A: ....
                qa = f"{num}Q: {q}\nA: {a}"
            else:
                # Hanya jawaban, nomor (jika dipilih) di depan
                qa = f"{num}{a}"

            items.append(qa)

        sep = "\n\n" if blank_between else "\n"
        body = sep.join([s for s in items if s])

        header = f"**{sesi_pick}**\n" if include_header else ""
        wa_text = (header + body).strip()
        encoded_txt = quote(wa_text, safe="")   # encode SEMUA karakter

        # Preview
        st.markdown("#### Preview teks WA")
        st.code(wa_text, language=None)

        # --- Hitung panjang teks ---
        encoded = quote(wa_text)
        len_encoded = len(encoded)

        # --- Batas URL aman ---
        URL_LIMIT = 2000

        st.subheader("Kirim Pesan ke WhatsApp")

        if len_encoded <= URL_LIMIT:
            wa_url = f"https://web.whatsapp.com/send?text={encoded}"

            components.html(f"""
<button id="wa-btn"
style="padding:.5rem .75rem;border:1px solid #d0d7de;border-radius:8px;
background:white;cursor:pointer">
💬 Kirim lewat WhatsApp
</button>
<script>
const url = "{wa_url}";
document.getElementById("wa-btn").onclick = () => {{
window.open(url, "waweb");
}};
</script>
""", height=60)

            st.success(f"Teks aman untuk dikirim (panjang encoded: {len_encoded} karakter).")
        else:
            st.error(f"Teks terlalu panjang untuk dikirim langsung ke WhatsApp "
                     f"(panjang encoded: {len_encoded} karakter, batas: {URL_LIMIT}).")
            st.info("👉 Sebaiknya copy-paste teks ini langsung di WhatsApp Web.")

# --- Tombol Copy + Download sejajar pakai flexbox ---
            components.html(f"""
<div style="display:flex; gap:10px; margin-top:10px;">

<!-- Tombol Copy -->
<button id="copy-btn" style="
padding:0.6rem 1rem;
background-color:#25D366;
color:white;
border:none;
border-radius:6px;
cursor:pointer;
font-size:14px;
">📋 Copy Teks</button>

<!-- Tombol Download (dibuat tombol HTML biasa) -->
<a id="download-btn"
download="jawaban_{{sesi_pick.replace(' ', '_')}}.txt"
href="data:text/plain;charset=utf-8,{{encoded_txt}}"
style="
padding:0.6rem 1rem;
background-color:#0d6efd;
color:white;
border:none;
border-radius:6px;
cursor:pointer;
font-size:14px;
text-decoration:none;">
⬇️ Download .txt
</a>

</div>

<script>
const btn = document.getElementById("copy-btn");
btn.onclick = function() {{
navigator.clipboard.writeText(`{wa_text}`);
btn.innerText = "✅ Teks sudah dicopy!";
setTimeout(()=>btn.innerText="📋 Copy Teks", 2000);
}};
</script>
""", height=100)
   
   
    else:
        st.info("Pilih satu sesi wawancara dulu untuk menyiapkan teks WA.")


    # (Opsional) intip 5 baris pertama untuk memastikan kolom 'sesi' ada
    #st.write("Kolom di df_view:", list(df_view.columns))
    #st.write("Preview df_view:", df_view.head())

    # ---------- Konfigurasi Ag-Grid ----------
    gb = GridOptionsBuilder.from_dataframe(df_view)

    # Kolom grouping (sembunyikan di anak, tampil sebagai header grup)
    gb.configure_column(
        "sesi",
        header_name="Tanggal | Jam | Media | Wartawan",
        rowGroup=True,            # <— inilah yang bikin grouping
        hide=True                 # sembunyikan dari baris anak
    )

    # Kolom No di kiri
    gb.configure_column(
        "No", header_name="No", width=70, pinned="left",
        sortable=False, filter=False
    )

    # ==== style anti-hyphenation (opsional tapi bagus) ====
    _common_cell_style = {
        "hyphens": "none", "wordBreak": "normal",
        "overflowWrap": "break-word", "lineHeight": "1.45",
    }

    # Kolom lain
    gb.configure_column("id", header_name="ID", hide=True)
    gb.configure_column(
        "question", header_name="pertanyaan",
        wrapText=True, autoHeight=True, minWidth=420, flex=3, cellStyle=_common_cell_style
    )
    gb.configure_column(
        "answer", header_name="jawaban",
        wrapText=True, autoHeight=True, minWidth=520, flex=4, cellStyle=_common_cell_style
    )

    # Opsi grouping
    gb.configure_grid_options(
        domLayout="normal",
        rowModelType="clientSide",

        # 🔹 tampilkan baris grup sebagai header yang bisa dibuka
        groupDisplayType="multipleColumns",     # <— lebih stabil di versi 1.1.x
        groupDefaultExpanded=0,                 # kolaps dulu biar cepat
        groupMaintainOrder=True,
        groupHideOpenParents=False,
        showOpenedGroup=True,                   # <— penting agar bisa dibuka manual
        suppressGroupRowsSticky=False,

        # 🔹 pengaturan umum
        rowSelection="multiple",
        groupSelectsChildren=True,
        suppressRowClickSelection=True,

        # 🔹 pagination tetap aktif
        pagination=True,
        paginationPageSize=20,
        paginateChildRows=True,
        suppressPaginationPanel=False,

        # 🔹 definisi kolom grup di kiri
        autoGroupColumnDef={
            "headerName": "Tanggal | Jam | Media | Wartawan",
            "minWidth": 420,
            "pinned": "left",  # biar gak hilang
            "cellRendererParams": {"suppressCount": False},
            "cellStyle": {
                "display": "flex",
                "alignItems": "center",
                "justifyContent": "center",
                "textAlign": "center",
                "fontWeight": "600",
            },
            "checkbox": True,
        },
    )

    # Checkbox di baris anak juga aktif
    gb.configure_selection(selection_mode="multiple", use_checkbox=True)

    # Pagination (C) dan quick search
    gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
    gb.configure_grid_options(quickFilter=True)

    grid_options = gb.build()

    # --- CSS jaga-jaga biar footer tidak ketutup ---
    st.markdown("""
    <style>
    .ag-theme-streamlit .ag-paging-panel { display:flex !important; visibility:visible !important; opacity:1 !important; }
    .ag-theme-streamlit .ag-root-wrapper { height: 100% !important; }
    .ag-theme-streamlit .ag-root-wrapper-body { height: calc(100% - 34px) !important; }
    </style>
    """, unsafe_allow_html=True)

    # ---------- Render grid ----------
    grid_response = AgGrid(
        df_view,
        gridOptions=grid_options,
        columns_auto_size_mode=ColumnsAutoSizeMode.NO_AUTOSIZE,
        fit_columns_on_grid_load=True,
        enable_enterprise_modules=True,
        allow_unsafe_jscode=False,
        update_mode=GridUpdateMode.SELECTION_CHANGED,   # penting agar ceklis memicu rerun
        height=560,
    )

    # 2) Ambil raw selection
    sel_raw = grid_response.get("selected_rows", [])

    # Kadang plugin bisa mengembalikan list[dict] atau DataFrame; samakan jadi DataFrame
    if isinstance(sel_raw, pd.DataFrame):
        sel_df = sel_raw.copy()
    else:
        sel_df = pd.DataFrame(sel_raw)

    # 3) Ekstrak kolom id (abaikan node grup yang tidak punya id)
    s_id = sel_df["id"] if "id" in sel_df.columns else pd.Series([], dtype="object")

    # 4) Bersihkan & konversi jadi integer list
    #    - buang None/NaN
    #    - ambil hanya digit kalau ada noise
    #    - konversi ke int
    ids_to_delete = (
        s_id.dropna()
            .astype(str)
            .str.extract(r"(\d+)")[0]   # ambil digit saja
            .dropna()
            .astype(int)
            .tolist()
    )

    # --- Tombol hapus ---
    col_del, col_hint = st.columns([1, 3])
    with col_del:
        if st.button(f"🗑️ Hapus terpilih ({len(ids_to_delete)})", key=f"del_grp"):
            if not ids_to_delete:
                st.warning("Belum ada baris (leaf) yang dicentang.")
            else:
                ok = delete_rows_by_ids(ids_to_delete)
                if ok:
                    st.success(f"Berhasil menghapus {len(ids_to_delete)} baris.")
                    st.rerun()

    with col_hint:
        st.caption("Header sesi menampilkan Tanggal | Jam | Media | Wartawan. "
                   "Centang **leaf** (baris Q&A) di bawahnya untuk memilih.")

    # >>> INSERT D — Panel Edit Baris Terpilih (muncul di Tab Rekap, di bawah tabel)
    st.markdown("---")
    st.subheader("✏️ Edit baris terpilih")

    # Gunakan selection yang sudah kamu punya: sel_raw
    # Batasi 1 baris agar aman
    selected_count = len(sel_raw) if isinstance(sel_raw, list) else (len(sel_raw) if isinstance(sel_raw, pd.DataFrame) else 0)

    if selected_count == 0:
        st.caption("Centang **1** baris Q&A di tabel untuk mulai mengedit.")
    elif selected_count > 1:
        st.info("Pilih **satu** baris saja untuk diedit.")
    else:
        # Normalisasi baris terpilih ke dict
        if isinstance(sel_raw, list):
            row = sel_raw[0]
        else:
            row = sel_raw.iloc[0].to_dict()

        try:
            row_id = int(str(row.get("id", "")).strip())
        except Exception:
            row_id = 0

        q_init = (row.get("question") or "").strip()
        a_init = (row.get("answer") or "").strip()

        with st.form(key=f"edit_form_{row_id}", clear_on_submit=False):
            st.caption(f"ID: {row_id}")
            new_q = st.text_area("Pertanyaan", value=q_init, height=140)
            new_a = st.text_area("Jawaban", value=a_init, height=160)

            col_save, col_help = st.columns([1, 3])
            with col_save:
                submitted = st.form_submit_button("💾 Simpan Perubahan")
            with col_help:
                st.caption("Periksa isian sebelum menyimpan.")

            if submitted:
                if not row_id:
                    st.warning("ID baris tidak valid.")
                else:
                    ok = update_row_by_id(row_id, new_q.strip(), new_a.strip())
                    if ok:
                        st.success("Berhasil menyimpan perubahan.")
                        st.rerun()
                    else:
                        st.error("Gagal menyimpan perubahan. Coba lagi.")

with tabs[1]:
    render_rekap()  # <- panggil fungsi di dalam tab
       
#st.write("checkpoint: sebelum EXPORT")
# =============== TAB EXPORT (Lazy Load) ===============
with tabs[2]:
    st.subheader("⬇️ Export Data")

    # ------- FUNGSI DIDEFINISIKAN DI ATAS, SEBELUM DIPANGGIL -------
    import time
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        st.error("Pustaka 'openpyxl' belum terpasang. Jalankan: pip install openpyxl")
        st.stop()

    def safe_query_chunked(limit=5000, max_rows=200000):
        """Ambil data bertahap dari Supabase agar tidak timeout."""
        all_chunks, offset, total = [], 0, 0
        pbar = st.progress(0.0)
        while True:
            try:
                df = query_questions(qkey="", media="", journalist="",
                                     limit=limit, offset=offset)
            except Exception as e:
                st.warning(f"Gagal ambil batch offset={offset}: {e}")
                break
            if df is None or df.empty:
                break
            all_chunks.append(df)
            n = len(df); total += n; offset += n
            pbar.progress(min(1.0, total / max_rows))
            if n < limit or total >= max_rows:
                break
            time.sleep(0.3)
        pbar.empty()
        if not all_chunks:
            return pd.DataFrame()
        return pd.concat(all_chunks, ignore_index=True)

    # ------- UI TRIGGER -------
    # HANYA tombol ini; jangan pakai "or st.session_state.get('export_loaded')"
    if st.button("📦 Muat & Siapkan File Excel", key="btn_load_export"):
        st.session_state["export_ready"] = False

        # Ambil data
        with st.spinner("Mengambil data dari Supabase..."):
            df = safe_query_chunked(limit=5000)

        if df.empty:
            st.info("Belum ada data untuk diekspor.")
            st.stop()

        # Pastikan kolom penting ada
        for c in ("id", "question", "answer", "session_time", "media", "journalist"):
            if c not in df.columns:
                df[c] = ""

        ts = pd.to_datetime(df.get("session_time"), errors="coerce")
        df["tanggal"] = ts.dt.strftime("%Y-%m-%d").fillna("")
        df["jam"] = ts.dt.strftime("%H:%M").fillna("")
        df["sesi"] = df["tanggal"] + " | " + df["jam"] + " | " + df["media"] + " | " + df["journalist"]
        df = df.sort_values(["session_time", "id"], ascending=[False, False])

        # Bangun workbook
        wb = Workbook(); ws = wb.active; ws.title = "Rekap"
        headers = ["Sesi", "ID", "Pertanyaan", "Jawaban"]; ws.append(headers)
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        row_idx = 2
        for sesi, g in df.groupby("sesi", sort=False):
            start_row = row_idx
            for _, r in g.iterrows():
                ws.cell(row=row_idx, column=1, value=sesi)
                ws.cell(row=row_idx, column=2, value=int(r["id"]) if pd.notnull(r["id"]) else None)
                ws.cell(row=row_idx, column=3, value=str(r.get("question","")))
                ws.cell(row=row_idx, column=4, value=str(r.get("answer","")))
                for col in (3,4):
                    ws.cell(row=row_idx, column=col).alignment = Alignment(wrapText=True, vertical="top")
                row_idx += 1
            if row_idx - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row_idx-1, end_column=1)
                mc = ws.cell(row=start_row, column=1)
                mc.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                mc.font = Font(bold=True)

        # border & width
        widths = (14, 4, 60, 95)
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                if cell.alignment is None:
                    cell.alignment = Alignment(vertical="top")

        # simpan ke session_state
        bio = BytesIO(); wb.save(bio); bio.seek(0)
        st.session_state["excel_data"] = bio.getvalue()
        st.session_state["export_ready"] = True
        st.success("✅ File Excel berhasil disiapkan. Silakan unduh di bawah.")

    # Tampilkan tombol download HANYA jika file sudah siap
    if st.session_state.get("export_ready"):
        st.download_button(
            "💾 Download Excel",
            data=st.session_state["excel_data"],
            file_name=f"rekap_siapcuaca_{pd.Timestamp.now():%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_download_export"
        )
